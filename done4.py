#-*-coding:GBK -*- 
#�������������ͳ�Ʊ�
import numpy as np
import pandas as pd
from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
pd.options.mode.chained_assignment = None

def cell_style(ws,len_index,names,names2):
	width_dict={'A':44,'B':9,'C':9,'D':14,'E':9}
	thin = Side(border_style='thin',color='00000000')
	alignment = Alignment(horizontal='center',vertical='center',wrap_text=False)
	alignment1 = Alignment(horizontal='center',vertical='center',wrap_text=True)
	font = Font(name='����',size=20,bold=True)
	font1 = Font(bold=True)
	font2 = Font(bold=True,color='FFFF0000')

	ws['A1'] =str(names)+'>>'+names2+'�̲�������ͳ��'+'('+str(len_index)+'�ְ���)'
	ws['A2'] = 'У������'+names2+'�༶����'
	ws['B2'] = '����'+names2+'����������'
	ws['C2'] = '��ֹĿǰ�ѽɷ�����'
	ws['D2'] = names2+'����ȫ���̲�������(�ÿ�Ŀ���а���޶����)'
	ws['E2'] = '����̲�(D�м�ȥC��)'
	ws.merge_cells('A1:E1')

	ws['A1'].font = font
	# ~ ws.row_dimensions[2].height = 30
	for row in ws.iter_rows(min_row=1,max_row=len_index+2,min_col=1,max_col=5):
		for cell in row:
			cell.border = Border(top=thin,left=thin,right=thin,bottom=thin)
			cell.alignment = alignment
	for c in ws[2]:
		c.alignment = alignment1
		c.font = font1
	ws.row_dimensions[2].height = 70
	
	for k,v in width_dict.items():
		ws.column_dimensions[k].width = v

	for i in range(3,len_index+3):
		ws['E'+str(i)].value=ws['D'+str(i)].value - ws['C'+str(i)].value
		if int(ws['E'+str(i)].value)<=0:
			ws['E'+str(i)].font = font2
	for i in range(3,len_index+2):
		ws.row_dimensions[i].height = 14


def wash_data(filename):

	data = pd.read_excel(filename, sheet_name='Sheet0',usecols=[4,6,7,9,11,17,18,22,24,29,31])#��ȡ��
	finish_excel = data.loc[2,'��ѧ��']+ '__'+ data.loc[2,'ѧ��']
	data.rename(columns = {'ѧ��':'�༶����'},inplace=True)
	data.��� = data.���.str.replace('^[\u4e00-\u9fa5][\u4e00-\u9fa5][\u4e00-\u9fa5]|˫ʦ','')
	
	class_dic = {'Сѧһ':'<1>һ','Сѧ��':'<2>��','Сѧ��':'<3>��','Сѧ��':'<4>��','Сѧ��':'<5>��','Сѧ��':'<6>��',
				'����һ':'<7>��һ','���ж�':'<8>����','������':'<9>����'}
	for k,v in class_dic.items():
		data.�꼶 = data['�꼶'].str.replace(k,v)
		
	data.�༶���� = data['�༶����'].str.cat(data['�꼶'],join='left',sep=' ')
	data.�༶���� = data['�༶����'].str.cat(data['ѧ��'],join='left',sep=' ')
	data.�༶���� = data['�༶����'].str.cat(data['���'],join='left',sep=' ')
	data.��� = data.���.str.replace('^[\u4e00-\u9fa5][\u4e00-\u9fa5][\u4e00-\u9fa5]','')
	
	
	data = data.drop(columns=['�꼶','���','��ʦ','����','��ѧ��','�Ͽ�ʱ��','�ܿδ�'])
	banshu  = data.�༶����.value_counts()
	banshu = pd.DataFrame(banshu)

	tongji = data.groupby('�༶����').sum()

	frames = [banshu,tongji]
	result = pd.concat(frames,axis=1,sort=False)

	
	course = ['��ѧ','Ӣ��','����','����','��ѧ','��ѧ']
	wb = Workbook()
	for c in course:
		ws = wb.create_sheet(c,-1)
		result1 = result.loc[result.index.str.contains(c)]
		result1 = result1.sort_index()
		for r in dataframe_to_rows(result1,index=True,header=True):
			ws.append(r)
		cell_style(ws,len(result1.index),finish_excel,c)
	finish_excel = finish_excel+'__'+'�������ͳ�Ʊ�.xlsx'
	wb.save(finish_excel)
	return str(finish_excel)


# ~ filename = '���´���.xlsx'
# ~ wash_data(filename)










































