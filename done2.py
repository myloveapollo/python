#-*-coding:GBK -*- 
#ǰ̨�α�
import numpy as np#����nump����ģ��
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
pd.options.mode.chained_assignment = None


def cell_style0(ws0,len_index0,week0,names):
	width_dict={'A':6.7,'B':11,'C':5,'D':20,'E':16,'F':15,'G':11,'H':11,'I':21,'J':7.5,'K':5}
	thin = Side(border_style='thin',color='00000000')
	alignment = Alignment(horizontal='center',vertical='center',wrap_text=False)
	font = Font(name='����',size=20,bold=True)
	font1 = Font(bold=True)
	font2 = Font(bold=False,color='FFFF0000')
	
	ws0.insert_rows(1)
	ws0['A1']= names+' '+week0+'�α�('+str(len_index0)+'����)'
	ws0.merge_cells('A1:K1')
	ws0['A1'].font = font
	ws0.row_dimensions[1].height = 30
	
	for row in ws0.iter_rows(min_row=1,max_row=len_index0+2,min_col=1,max_col=11):
		for cell in row:
			if cell.value == datetime.datetime.now().strftime('%Y-%m-%d'):#����뵱ǰ������ȵĽ��ʱ��
				cell.font = font2
			cell.border = Border(top=thin,left=thin,right=thin,bottom=thin)
			cell.alignment = alignment
	for c in ws0[2]:
		c.font = font1
	for k,v in width_dict.items():
		ws0.column_dimensions[k].width = v
	for i in range(2,len_index0+3):
		ws0.row_dimensions[i].height = 16



def wash_data(filename):
	wb = Workbook()
	data = pd.read_excel(filename, sheet_name='Sheet0',usecols=[4,6,7,9,11,14,17,18,20,21,22,24,29])#��ȡ��
	
	weekdays_all = ['����','��������','��������','��������','��������'
					,'��������','��������','��������','��������','�ܶ�','����','����']
					
	weekdays_all2 = ['����','����','����','����']	
	finish_excel = data.loc[2,'��ѧ��']+ '__'+ data.loc[2,'ѧ��']
	data.��ʦ =  data.��ʦ.str.replace('[0-9]\d*$','')#ɾ����ʦ���ƺ������

	data_fudao = data.������ʦ.str.replace('[0-9]\d*$','')#ɾ��������ʦ���ƺ������
	data_fudao2 = data_fudao.str.replace('.*[\u4e00-\u9fa5]','����:',regex=True)#ɾ�����ֺ�ֻʣ���֣�����ȫ���滻�ɸ���
	data.������ʦ = data_fudao2 + data_fudao #����+��ʦ����
	
	
	data.��ʦ = data.��ʦ.str.cat(data.������ʦ,join='left',sep=' ')#��ʦ�븨����ʦ�϶�Ϊһ
	data.���� = data.����.str.replace('[(].*?[)]|[��].*?[��]','',regex=True)
	# ~ data.��� = data.���.str.replace('^[\u4e00-\u9fa5][\u4e00-\u9fa5][\u4e00-\u9fa5]','')
	data.rename(columns = {'�ѽ�����':'����'},inplace=True)	#'���Ͽδ�':'�δ�',
	data.sort_values(by='�Ͽ�ʱ��', axis=0, ascending=True,inplace=True)#���Ͽ�ʱ�����򷽷�
	data = data.drop(columns=['������ʦ','��ѧ��'],axis=1)


	if data.ѧ��.iloc[1] == '������' or data.ѧ��.iloc[1] == '�＾��':
		for week in weekdays_all:
			data_num = data.loc[data.�Ͽ�ʱ��.str.contains(week)]
			if week =='�ܶ�'or week =='����' or week =='����' or week=='����':
				data_num.sort_values(by='����',axis=0,ascending=True,inplace=True)
			else:
				data_num.sort_values(by=['�Ͽ�ʱ��','����'],axis=0,ascending=True,inplace=True)
			ws = wb.create_sheet(week+'('+str(len(data_num.index))+'����)',-1)
			for r in dataframe_to_rows(data_num,index=False,header=True):
				ws.append(r)
			cell_style0(ws,len(data_num.index),week,finish_excel)
			
	elif data.ѧ��.iloc[1] == '���ڰ�' or data.ѧ��.iloc[1] == '���' or data.ѧ��.iloc[1] == '�����':
		week = data.ѧ��.iloc[1]
		ws = wb.create_sheet(week,-1)
		data.sort_values(by='�������',axis=0,ascending=True, inplace=True)#�������������
		for r in dataframe_to_rows(data,index=False, header=True):
			ws.append(r)
		cell_style0(ws,len(data.index),week,finish_excel)
	
	else:#if data.ѧ��.iloc[1] == '��ٰ�' or data.ѧ��.iloc[1] == '���ٰ�'
		data['�Ͽ�ʱ��'] = data['�Ͽ�ʱ��'].str.replace('һ','1')
		data['�Ͽ�ʱ��'] = data['�Ͽ�ʱ��'].str.replace('��','2')
		data['�Ͽ�ʱ��'] = data['�Ͽ�ʱ��'].str.replace('��','3')
		data['�Ͽ�ʱ��'] = data['�Ͽ�ʱ��'].str.replace('��','4')
		data['�Ͽ�ʱ��'] = data['�Ͽ�ʱ��'].str.replace('��','0')
		for week in weekdays_all2:
			data_num = data.loc[data.�Ͽ�ʱ��.str.contains(week)]
			data_num.sort_values(by=['�Ͽ�ʱ��','����'],axis=0,ascending=True,inplace=True)
			ws = wb.create_sheet(week+'('+str(len(data_num.index))+'����)',-1)
			for r in dataframe_to_rows(data_num,index=False,header=True):
				ws.append(r)
			cell_style0(ws,len(data_num.index),week,finish_excel)
	finish_excel = finish_excel +'__'+'ǰ̨�α�.xlsx'
	wb.save(finish_excel)
	return str(finish_excel)

# ~ filename = '���¶��ڰ�.xlsx'
# ~ wash_data(filename)








































