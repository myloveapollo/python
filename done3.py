#-*-coding:GBK -*-
#������������ķ�����
import numpy as np 
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
pd.options.mode.chained_assignment = None


def cell_style(ws,len_index):
	width_dict = {'B':30,'C':6,'D':14.56,'E':12.00,'F':12.00,'G':6,'H':6}
	font = Font(name='����',size=20,bold=True)
	thin = Side(border_style='thin',color='00000000')
	alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
	for row in ws.iter_rows(min_row=1,max_row=len_index,min_col=2,max_col=8):
		for cell in row:
			cell.font=font
			cell.border = Border(bottom=thin)
			cell.alignment = alignment
	for k,v in width_dict.items():
		ws.column_dimensions[k].width = v
	for i in range(len_index+1):
		ws.row_dimensions[i].height = 165
	
def wash_data(filename):
	wb = Workbook()
	data = pd.read_excel(filename, sheet_name='Sheet0',usecols=[4,6,7,9,11,14,17,18,22,25,29])#��ȡ��
	data['���Ͽδ�'] = data['���Ͽδ�']+1

	weekdays_all = ['����','��������','��������','��������','��������'
					,'��������','��������','��������','��������','�ܶ�','����','����']
					
	weekdays_all2 = ['����','����','����','����']	
	
	finish_excel = data.loc[2,'��ѧ��']+ '__'+ data.loc[2,'ѧ��']
	
	data_fudao = data.������ʦ.str.replace('[0-9]\d*$','')#ɾ��������ʦ���ƺ������
	data_fudao2 = data_fudao.str.replace('.*[\u4e00-\u9fa5]','����',regex=True)#ɾ�����ֺ�ֻʣ���֣�����ȫ���滻�ɸ���
	data.������ʦ = data_fudao2 +'\n'+ data_fudao #����+��ʦ����
	
	data.��ʦ =  data.��ʦ.str.replace('[0-9]\d*$','')#ɾ����ʦ���ƺ������
	data.��ʦ = data.��ʦ + '\n'+data.������ʦ
	# ~ data.��ʦ = data['��ʦ'].str.cat(data['������ʦ'],join='left',sep=' ')
	data.�꼶 = data['�꼶'].str.cat(data['���'],join='left')
	data.���� = data.����.str.replace('[(].*?[)]|[��].*?[��]','')

	data.rename(columns = {'���Ͽδ�':'�δ�','�ѽ�����':'����'},inplace=True)	
	data.sort_values(by='�Ͽ�ʱ��', axis=0, ascending=True,inplace=True)#���Ͽ�ʱ�����򷽷�
	data = data.drop(columns=['������ʦ'],axis=1)
	data = data.drop(columns=['���','��ѧ��'],axis=1)
	# ~ data = data.drop(columns=['�δ�'],axis=1)

	if data.ѧ��.iloc[1] == '������' or data.ѧ��.iloc[1] == '�＾��':
		for week in weekdays_all:
			data_num = data.loc[data.�Ͽ�ʱ��.str.contains(week)]
			if week =='�ܶ�'or week =='����' or week =='����' or week=='����':
				data_num.sort_values(by='����',axis=0,ascending=True,inplace=True)
			ws =wb.create_sheet(week+'('+str(len(data_num.index))+'����)',-1)
			ws.column_dimensions.group('A',hidden=True)
			for r in dataframe_to_rows(data_num,index=False,header=False):
				ws.append(r)
			cell_style(ws,len(data_num.index))
	else:
		for week in weekdays_all2:
			data_num = data.loc[data.�Ͽ�ʱ��.str.contains(week)]
			ws = wb.create_sheet(week+'('+str(len(data_num.index))+'����)',-1)
			ws.column_dimensions.group('A',hidden=True)
			for r in dataframe_to_rows(data_num,index=False,header=False):
				ws.append(r)
			cell_style(ws,len(data_num.index))
	finish_excel = finish_excel+'__'+'��ķ�����.xlsx'
	wb.save(finish_excel)
	return str(finish_excel)
	

# ~ filename = '�������4.21.xlsx'
# ~ wash_data(filename)









































