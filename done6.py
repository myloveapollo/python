#-*-coding:GBK -*- 
#��Ӫ�����Űർ���
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font,numbers
from openpyxl.worksheet.datavalidation import DataValidation
pd.options.mode.chained_assignment = None

def cell_style(ws,len_index):
	width_dict={'A':15.13,'B':15.13,'C':19.28,'D':15.13,'E':15.13,'F':15.13}
	thin = Side(border_style=None,color='00000000')
	alignment = Alignment(horizontal='center',vertical='center',wrap_text=False)
	font = Font(name='����',size=11,bold=False)
	
	list_content = ('"06:00,06:30,07:00,07:30,08:00,08:30,09:00,09:30,10:00,10:30,11:00,11:30,'
					'12:00,12:30,13:00,13:30,14:00,14:30,15:00,15:30,16:00,16:30,17:00,17:30,'
					'18:00,18:30,19:00,19:30,20:00,20:30,21:00,21:30,22:00,22:30,23:00,23:30"')#00:00,00:30,01:00,01:30,02:00,02:30,03:00,03:30,04:00,04:30,05:00,05:30,'

	dv_type = DataValidation(type="list", formula1='"���,����,��Ϣ,����ְȱ��,��ѵ,����,ҽ����,�¼�,���,����,����,�����,ɥ��"', allow_blank=False)
	dv_time = DataValidation(type="list", formula1=list_content,operator='equal', allow_blank=True)
	ws.add_data_validation(dv_type)
	ws.add_data_validation(dv_time)
	type_index = 'D2:D'+str(len_index+1)
	time_index = 'E2:F'+str(len_index+1)
	dv_type.add(type_index)
	dv_time.add(time_index)
	
	for row in ws.iter_rows(min_row=2,max_row=len_index+1,min_col=3,max_col=3):
		for cell in row:
			cell.number_format = 'yyyy-mm-dd'
			
	for row in ws.iter_rows(min_row=2,max_row=len_index+1,min_col=5,max_col=6):
		for cell in row:
			cell.number_format = 'hh:mm'
	
	for row in ws.iter_rows(min_row=1,max_row=len_index+1,min_col=1,max_col=6):
		for cell in row:
			cell.border = Border(top=thin,left=thin,right=thin,bottom=thin)
			cell.alignment = alignment
			cell.font = font

	for k,v in width_dict.items():
		ws.column_dimensions[k].width = v


def ceshi(data):
	writer = pd.ExcelWriter('ceshi.xlsx')
	data.to_excel(writer,index=True,header=True)
	writer.save()


def wash_data(filename):
	names=['����','����','��һ','�ܶ�','����','����','����','����','����']
	data = pd.read_excel(filename,sheet_name=0,header=None,names=names,usecols=[1,2,3,5,7,9,11,13,15])#��ȡ��,

	data_time = data.iloc[1]#��ȡ�����ѱ���,��ʽҪ������
	name_a = str(data.iloc[1,2])[6:10]+'��'+str(data.iloc[1,8])[6:10]
	in_excel = name_a + '��Ӫ�����Űർ��.xlsx' #����������ɵı� 
	
	# ~ data.drop(axis=1,columns=[0],inplace=True)
	data.drop(axis=0,index=[1,0,2,3,4],inplace=True)#ɾ��û�е�����Ϣ
	data.dropna(axis=0,how='all',inplace=True)#ɾ��ȫ�յ�����Ϣ
	data.replace('OFF','OFF-OFF',regex=True,inplace=True)#OFF����ڿ� ������ʽ�ӣ������ִ�Сд
	data = data.T#����

	#���һ����
	data_make = pd.DataFrame(np.full([7*len(data.columns),6],np.nan),columns=['����','����','����(YYYY-MM-DD)','����','�ϰ�ʱ��','�°�ʱ��'])
	
	data_job_num = []
	for nam in list(data.iloc[0]):
		for i in range(1,8):
			data_job_num.append(nam)
	data_make['����'] = data_job_num

	#��ȡ����
	data_col_name = []
	for nam in list(data.iloc[1]):
		for i in range(1,8):
			data_col_name.append(nam)
	data_make['����']= pd.Series(data_col_name)#��������

	#��ȡʱ��
	data.drop(axis=0,index='����',inplace=True)
	data.drop(axis=0,index='����',inplace=True)
	data_col_arrivetime = []#�ϰ�ʱ��
	data_col_leavetime = []#�°�ʱ��
	data_type =[]#����

	for col in list(data.columns):
		data_split_f = data[col].str.split('-').str[0]#�ϰ�ʱ����ǰ���
		data_split_s = data[col].str.split('-').str[1]#�°�ʱ���ú����
		for f in list(data_split_f):
			if len(str(f))==4:#len(f) == 4:
				f = '0'+f
			data_col_arrivetime.append(f)
		for s in list(data_split_s):
			data_col_leavetime.append(s)
			if s !='OFF':
				data_type.append('����')
			else:
				data_type.append('��Ϣ')

	data_make['����'] = data_type#������������
	data_make['�ϰ�ʱ��'] = data_col_arrivetime#�����ϰ�ʱ��
	data_make['�°�ʱ��'] = data_col_leavetime#�����°�ʱ��
	data_make.replace('OFF|off','',regex=True,inplace=True)
	
	# ~ #��ȡ'����(YYYY-MM-DD)'--------����1
	# ~ data_time = list(map(lambda x: '2019-'+x, data_time))#�ù�ʽ��ÿ���ַ�ǰ��2019-
	
	data_yyyy= []
	for data_t in range(0,len(list(data.columns))):
		for data_t in list(data_time[2:]):
			data_yyyy.append(str(data_t)[:-9])#��ÿ��col��datatime�����
	data_make['����(YYYY-MM-DD)'] = data_yyyy #��ֵ��col����
	
	#��openpyxl���ø�ʽ
	wb = Workbook()
	ws = wb.create_sheet('Sheet1',-1)
	for r in dataframe_to_rows(data_make,index=False,header=True):
		ws.append(r)
	cell_style(ws,len(data_make.index))
	wb.save(in_excel)
	return in_excel
	
# ~ filename = '�ƽ���.xlsx'
# ~ wash_data(filename)


